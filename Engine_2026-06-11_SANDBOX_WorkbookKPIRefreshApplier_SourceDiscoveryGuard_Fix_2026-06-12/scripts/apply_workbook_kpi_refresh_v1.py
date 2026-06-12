#!/usr/bin/env python3
"""
F1 Workbook/KPI Refresh Applier v1

Purpose:
- Convert Session Data Processor outputs into dated sandbox workbook/KPI refresh artifacts.
- Never overwrite the canonical workbook.
- Never alter Engine_2026-06-07_STABLE.
- Never promote a model layer.
- Never delete old files.

This script is repo-native and GitHub Actions safe. It writes only to:
- latest/workbook_kpi_refresh_applier/
- history/workbook_kpi_refresh_applier/<run_id>/

If a source workbook is provided, the script copies it first and appends sandbox sheets to the copy.
If not provided, it creates a lightweight sandbox workbook artifact containing readiness and audit tabs.
"""
from __future__ import annotations

import argparse
import csv
import datetime as dt
import hashlib
import json
import os
from pathlib import Path
import shutil
import sys
from typing import Any, Dict, List, Tuple

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except Exception as exc:  # pragma: no cover
    print(f"ERROR: missing spreadsheet dependency: {exc}", file=sys.stderr)
    sys.exit(22)

PROTECTED_STABLE_ENGINE = "Engine_2026-06-07_STABLE"

def utc_now() -> str:
    return dt.datetime.now(dt.timezone.utc).strftime("%Y%m%dT%H%M%SZ")

def read_json(path: Path, default: Any = None) -> Any:
    if not path.exists():
        return default
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)

def read_csv_rows(path: Path) -> List[Dict[str, Any]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))

def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2, sort_keys=True) + "\n", encoding="utf-8")

def write_csv(path: Path, rows: List[Dict[str, Any]], fieldnames: List[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if fieldnames is None:
        keys = []
        for row in rows:
            for k in row.keys():
                if k not in keys:
                    keys.append(k)
        fieldnames = keys or ["status"]
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for row in rows:
            w.writerow({k: row.get(k, "") for k in fieldnames})

def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()

def locate_session_processor_root(repo: Path) -> Path:
    """Locate the most recent concrete Session Data Processor output folder.

    The processor writes session artifacts one level or more below
    latest/session_data_processor/, for example:
      latest/session_data_processor/<event_slug>/<session_slug>/workbook_kpi_readiness.csv

    v1 looked only at latest/session_data_processor/ itself, which caused the
    applier to classify valid FP2 output as missing. This resolver searches
    recursively and chooses the newest folder containing the required source
    artifacts.
    """
    roots = [
        repo / "latest" / "session_data_processor",
        repo / "latest" / "session_data_processor_loop",
        repo / "latest" / "session_processor",
    ]
    required = {"sandbox_workbook_update_plan.json", "workbook_kpi_readiness.csv"}
    candidates: List[Path] = []
    for root in roots:
        if not root.exists():
            continue
        # Root itself may be a concrete output in some future configuration.
        if all((root / name).exists() for name in required):
            candidates.append(root)
        for plan in root.rglob("sandbox_workbook_update_plan.json"):
            folder = plan.parent
            if all((folder / name).exists() for name in required):
                candidates.append(folder)
    if candidates:
        # Prefer the newest concrete session output.
        return max(candidates, key=lambda x: max((f.stat().st_mtime for f in x.glob("*") if f.exists()), default=x.stat().st_mtime))
    return roots[0]

def has_required_session_sources(session_root: Path) -> Tuple[bool, List[str]]:
    required = [
        "sandbox_workbook_update_plan.json",
        "workbook_kpi_readiness.json",
        "workbook_kpi_readiness.csv",
        "forecast_bundle_ledger_snapshot.json",
    ]
    missing = []
    for name in required:
        path = session_root / name
        if not path.exists() or path.stat().st_size == 0:
            missing.append(name)
    # CSV must have at least header + one data row to be useful.
    csv_path = session_root / "workbook_kpi_readiness.csv"
    if csv_path.exists() and csv_path.stat().st_size > 0:
        try:
            rows = read_csv_rows(csv_path)
            if not rows:
                missing.append("workbook_kpi_readiness.csv:no_data_rows")
        except Exception as exc:
            missing.append(f"workbook_kpi_readiness.csv:unreadable:{exc}")
    return (len(missing) == 0), missing

def summarize_sources(readiness: Dict[str, Any], plan: Dict[str, Any], rows: List[Dict[str, Any]]) -> Tuple[str, bool, List[str]]:
    classification = (
        plan.get("classification")
        or readiness.get("classification")
        or readiness.get("overall_classification")
        or readiness.get("status")
        or "missing"
    )
    source_artifacts = []
    for key in ["sandbox_workbook_update_plan.json", "workbook_kpi_readiness.json", "workbook_kpi_readiness.csv", "forecast_bundle_ledger_snapshot.json"]:
        source_artifacts.append(key)
    material_change = classification not in ["missing"]
    if rows:
        material_change = True
    warnings = []
    if classification in ["needs_manual_review", "partial", "late", "conflicting"]:
        warnings.append(f"source_classification={classification}")
    if not rows:
        warnings.append("workbook_kpi_readiness_csv_missing_or_empty")
    return classification, material_change, warnings

def add_or_replace_sheet(wb, name: str):
    if name in wb.sheetnames:
        ws = wb[name]
        wb.remove(ws)
    return wb.create_sheet(name)

def style_header(ws, max_col: int):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1][:max_col]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

def write_table(ws, rows: List[Dict[str, Any]], preferred_fields: List[str] | None = None):
    if preferred_fields:
        fields = preferred_fields[:]
        for row in rows:
            for k in row.keys():
                if k not in fields:
                    fields.append(k)
    else:
        fields = []
        for row in rows:
            for k in row.keys():
                if k not in fields:
                    fields.append(k)
    if not fields:
        fields = ["status"]
        rows = [{"status": "no rows"}]
    ws.append(fields)
    for row in rows:
        ws.append([row.get(k, "") for k in fields])
    style_header(ws, len(fields))
    thin = Side(style="thin", color="D9E2F3")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(fields)):
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col_idx in range(1, len(fields)+1):
        width = min(max(12, max(len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, min(ws.max_row, 40)+1)) + 2), 42)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

def create_or_update_sandbox_workbook(source_workbook: Path | None, output_workbook: Path, readiness_rows: List[Dict[str, Any]], plan: Dict[str, Any], ledger: Dict[str, Any], manifest: Dict[str, Any]) -> None:
    output_workbook.parent.mkdir(parents=True, exist_ok=True)

    if source_workbook and source_workbook.exists():
        shutil.copy2(source_workbook, output_workbook)
        wb = load_workbook(output_workbook)
    else:
        wb = Workbook()
        default = wb.active
        default.title = "Sandbox Refresh Summary"

    # Remove default empty sheet if it is still blank and not our summary sheet.
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        ws = wb["Sheet"]
        if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
            wb.remove(ws)

    summary = add_or_replace_sheet(wb, "Sandbox Refresh Summary")
    summary_rows = [
        {"Field": "Run ID", "Value": manifest.get("run_id")},
        {"Field": "Created UTC", "Value": manifest.get("created_utc")},
        {"Field": "Source Status", "Value": manifest.get("source_status")},
        {"Field": "Material Change", "Value": str(manifest.get("material_change"))},
        {"Field": "Stable Engine Modified", "Value": "FALSE"},
        {"Field": "Canonical Workbook Overwrite", "Value": "FALSE"},
        {"Field": "Promotion Allowed", "Value": "FALSE"},
        {"Field": "Output Classification", "Value": "sandbox_only"},
    ]
    write_table(summary, summary_rows, ["Field", "Value"])

    ws = add_or_replace_sheet(wb, "KPI Readiness")
    write_table(ws, readiness_rows, ["source", "status", "rows", "classification", "session_key", "session_name", "updated_utc"])

    plan_ws = add_or_replace_sheet(wb, "Update Plan")
    plan_rows = []
    def flatten(prefix, obj):
        if isinstance(obj, dict):
            for k, v in obj.items():
                flatten(f"{prefix}.{k}" if prefix else k, v)
        elif isinstance(obj, list):
            plan_rows.append({"Path": prefix, "Value": f"list[{len(obj)}]"})
        else:
            plan_rows.append({"Path": prefix, "Value": obj})
    flatten("", plan)
    write_table(plan_ws, plan_rows, ["Path", "Value"])

    ledger_ws = add_or_replace_sheet(wb, "Forecast Bundle Ledger")
    ledger_rows = []
    if isinstance(ledger, dict):
        for k, v in ledger.items():
            if isinstance(v, (dict, list)):
                ledger_rows.append({"Field": k, "Value": json.dumps(v)[:1000]})
            else:
                ledger_rows.append({"Field": k, "Value": v})
    write_table(ledger_ws, ledger_rows, ["Field", "Value"])

    gov_ws = add_or_replace_sheet(wb, "Governor Gates")
    gov_rows = [
        {"Gate": "Promotion Gate", "Status": "BLOCKED", "Rule": "No promotion from sandbox refresh"},
        {"Gate": "Stable Separation Gate", "Status": "ENFORCED", "Rule": "No stable P1-P20 overwrite"},
        {"Gate": "Canonical Workbook Gate", "Status": "ENFORCED", "Rule": "Create dated sandbox copy only"},
        {"Gate": "Deletion/Cleanup Gate", "Status": "BLOCKED", "Rule": "No delete authority"},
        {"Gate": "Accuracy Objective Lock", "Status": "ENFORCED", "Rule": "No accuracy claim without live/blind validation"},
    ]
    write_table(gov_ws, gov_rows, ["Gate", "Status", "Rule"])

    audit_ws = add_or_replace_sheet(wb, "Refresh Audit")
    audit_rows = [{"Field": k, "Value": json.dumps(v) if isinstance(v, (dict, list)) else v} for k, v in manifest.items()]
    write_table(audit_ws, audit_rows, ["Field", "Value"])

    wb.save(output_workbook)

def make_report(manifest: Dict[str, Any], warnings: List[str]) -> str:
    status = "Pass with warnings" if warnings else "Pass"
    warnings_md = "\n".join([f"- {w}" for w in warnings]) if warnings else "- None"
    outputs_md = "\n".join([f"- `{o['path']}` ({o.get('sha256','')[:12]}...)" for o in manifest.get("outputs", [])])
    return f"""# F1 Workbook/KPI Refresh Applier Report

## Verdict

{status}

## Run

- Run ID: `{manifest.get('run_id')}`
- Created UTC: `{manifest.get('created_utc')}`
- Source status: `{manifest.get('source_status')}`
- Material change: `{manifest.get('material_change')}`

## Warnings

{warnings_md}

## Outputs

{outputs_md}

## Governance

- Canonical workbook overwrite: **blocked**
- Stable engine modification: **blocked**
- Model promotion: **blocked**
- Delete/cleanup authority: **blocked**

## Classification

This is a sandbox workbook/KPI refresh artifact. It may support readiness, forecast bundle discipline, and fantasy/race prediction state refresh, but it does not promote model logic.
"""

def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--repo-root", default=".", help="Repository root")
    ap.add_argument("--source-workbook", default="", help="Optional source workbook path to copy before appending sandbox sheets")
    ap.add_argument("--event-label", default="auto", help="Optional event label for output naming")
    ap.add_argument("--safe-test", action="store_true", help="Run preflight only and write no forecast-affecting outputs")
    args = ap.parse_args()

    repo = Path(args.repo_root).resolve()
    run_id = utc_now()
    latest_out = repo / "latest" / "workbook_kpi_refresh_applier"
    history_out = repo / "history" / "workbook_kpi_refresh_applier" / run_id
    session_root = locate_session_processor_root(repo)

    if args.safe_test:
        required_sources_present, missing_sources = has_required_session_sources(session_root)
        diag = {
            "status": "safe_test_pass",
            "run_id": run_id,
            "repo_root": str(repo),
            "discovered_session_processor_root": str(session_root.relative_to(repo)) if session_root.exists() and session_root.is_relative_to(repo) else str(session_root),
            "session_processor_root_exists": session_root.exists(),
            "required_session_sources_present": required_sources_present,
            "missing_session_sources": missing_sources,
            "canonical_workbook_overwrite": False,
            "stable_engine_modified": False,
            "promotion_allowed": False,
        }
        write_json(repo / "_runtime" / "workbook_kpi_refresh_applier_safe_test.json", diag)
        print(json.dumps(diag, indent=2))
        return 0

    required_sources_present, missing_sources = has_required_session_sources(session_root)
    if not required_sources_present:
        diag = {
            "status": "no_action",
            "reason": "session_processor_sources_missing_or_incomplete",
            "commit_allowed": False,
            "run_id": run_id,
            "repo_root": str(repo),
            "discovered_session_processor_root": str(session_root.relative_to(repo)) if session_root.exists() and session_root.is_relative_to(repo) else str(session_root),
            "session_processor_root_exists": session_root.exists(),
            "missing_session_sources": missing_sources,
            "canonical_workbook_overwrite": False,
            "stable_engine_modified": False,
            "promotion_allowed": False,
        }
        write_json(repo / "_runtime" / "workbook_kpi_refresh_status.json", diag)
        print(json.dumps(diag, indent=2))
        return 0

    plan = read_json(session_root / "sandbox_workbook_update_plan.json", {})
    readiness_json = read_json(session_root / "workbook_kpi_readiness.json", {})
    readiness_rows = read_csv_rows(session_root / "workbook_kpi_readiness.csv")
    ledger = read_json(session_root / "forecast_bundle_ledger_snapshot.json", {})
    latest_manifest = read_json(repo / "latest" / "latest_manifest.json", {})
    data_readiness = read_json(repo / "latest" / "data_readiness.json", {})
    combined_manifest = read_json(repo / "latest" / "combined_source_manifest.json", {})

    source_status, material_change, warnings = summarize_sources(readiness_json, plan, readiness_rows)

    event_label = args.event_label
    if event_label == "auto":
        event = plan.get("event", {}) if isinstance(plan, dict) else {}
        session = plan.get("session", {}) if isinstance(plan, dict) else {}
        race = str(event.get("race_name") or readiness_json.get("race_name") or "unknown_event").replace(" ", "_").replace("/", "_")
        sess = str(session.get("session_name") or readiness_json.get("session_name") or "session").replace(" ", "_").replace("/", "_")
        event_label = f"{race}_{sess}"

    workbook_name = f"F1_Workbook_KPI_SANDBOX_Refresh_{event_label}_{run_id}.xlsx"
    source_wb = Path(args.source_workbook).resolve() if args.source_workbook else None

    # Build outputs in history first.
    manifest = {
        "run_id": run_id,
        "created_utc": dt.datetime.now(dt.timezone.utc).isoformat(),
        "source_status": source_status,
        "source_processor_root": str(session_root.relative_to(repo)) if session_root.is_relative_to(repo) else str(session_root),
        "material_change": material_change,
        "warnings": warnings,
        "source_artifacts": [
            str((session_root / "sandbox_workbook_update_plan.json").relative_to(repo)) if (session_root / "sandbox_workbook_update_plan.json").exists() else "missing:sandbox_workbook_update_plan.json",
            str((session_root / "workbook_kpi_readiness.json").relative_to(repo)) if (session_root / "workbook_kpi_readiness.json").exists() else "missing:workbook_kpi_readiness.json",
            str((session_root / "workbook_kpi_readiness.csv").relative_to(repo)) if (session_root / "workbook_kpi_readiness.csv").exists() else "missing:workbook_kpi_readiness.csv",
            str((session_root / "forecast_bundle_ledger_snapshot.json").relative_to(repo)) if (session_root / "forecast_bundle_ledger_snapshot.json").exists() else "missing:forecast_bundle_ledger_snapshot.json",
        ],
        "governance": {
            "stable_engine": PROTECTED_STABLE_ENGINE,
            "stable_engine_modified": False,
            "canonical_workbook_overwrite": False,
            "promotion_allowed": False,
            "delete_authority": False,
        },
        "outputs": [],
        "commit_allowed": True,
    }

    history_out.mkdir(parents=True, exist_ok=True)
    latest_out.mkdir(parents=True, exist_ok=True)

    # Derived artifacts
    refreshed_rows = readiness_rows or [{
        "source": "session_processor",
        "status": source_status,
        "rows": "",
        "classification": source_status,
        "session_key": (plan.get("session", {}) or {}).get("session_key", ""),
        "session_name": (plan.get("session", {}) or {}).get("session_name", ""),
        "updated_utc": manifest["created_utc"],
    }]

    artifacts = {
        "sandbox_workbook_update_plan_applied.json": {
            "applied": True,
            "source_plan": plan,
            "source_status": source_status,
            "material_change": material_change,
            "governance": manifest["governance"],
        },
        "upstream_latest_manifest_snapshot.json": latest_manifest,
        "upstream_data_readiness_snapshot.json": data_readiness,
        "upstream_combined_source_manifest_snapshot.json": combined_manifest,
    }
    for name, obj in artifacts.items():
        write_json(history_out / name, obj)
        shutil.copy2(history_out / name, latest_out / name)

    write_csv(history_out / "refreshed_workbook_kpi_readiness.csv", refreshed_rows)
    shutil.copy2(history_out / "refreshed_workbook_kpi_readiness.csv", latest_out / "refreshed_workbook_kpi_readiness.csv")

    workbook_path = history_out / workbook_name
    create_or_update_sandbox_workbook(source_wb, workbook_path, refreshed_rows, plan, ledger, manifest)
    shutil.copy2(workbook_path, latest_out / workbook_name)

    # Manifest after outputs exist
    output_paths = [
        history_out / "sandbox_workbook_update_plan_applied.json",
        history_out / "refreshed_workbook_kpi_readiness.csv",
        workbook_path,
        history_out / "upstream_latest_manifest_snapshot.json",
        history_out / "upstream_data_readiness_snapshot.json",
        history_out / "upstream_combined_source_manifest_snapshot.json",
    ]
    for p in output_paths:
        manifest["outputs"].append({
            "path": str(p.relative_to(repo)) if p.is_relative_to(repo) else str(p),
            "sha256": sha256_file(p),
            "bytes": p.stat().st_size,
        })

    write_json(history_out / "workbook_kpi_refresh_manifest.json", manifest)
    shutil.copy2(history_out / "workbook_kpi_refresh_manifest.json", latest_out / "workbook_kpi_refresh_manifest.json")

    report = make_report(manifest, warnings)
    (history_out / "workbook_kpi_refresh_report.md").write_text(report, encoding="utf-8")
    shutil.copy2(history_out / "workbook_kpi_refresh_report.md", latest_out / "workbook_kpi_refresh_report.md")

    status_payload = {
        "status": "refresh_applied",
        "commit_allowed": True,
        "source_status": source_status,
        "material_change": material_change,
        "warnings": warnings,
        "latest_output_root": str(latest_out.relative_to(repo)) if latest_out.is_relative_to(repo) else str(latest_out),
        "history_output_root": str(history_out.relative_to(repo)) if history_out.is_relative_to(repo) else str(history_out),
        "sandbox_workbook": str((latest_out / workbook_name).relative_to(repo)) if (latest_out / workbook_name).is_relative_to(repo) else str(latest_out / workbook_name),
        "canonical_workbook_overwrite": False,
        "stable_engine_modified": False,
        "promotion_allowed": False,
    }
    write_json(repo / "_runtime" / "workbook_kpi_refresh_status.json", status_payload)

    print(json.dumps(status_payload, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
